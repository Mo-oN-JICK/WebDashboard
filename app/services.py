from __future__ import annotations

import io
import csv
from datetime import date, datetime, timedelta
from typing import Any

from flask import Response
from openpyxl import Workbook, load_workbook

from . import db
from .models import AuditLog, DailyMeasurement, ProcessMaster

METRICS = ["총체결", "NG", "NG율", "분류실패", "분류실패율", "Cluster"]


def parse_date(value: Any) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if value is None or str(value).strip() == "":
        raise ValueError("날짜 형식 오류")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError("날짜 형식 오류")


def to_int(value: Any, field: str) -> int:
    if value is None or value == "":
        return 0
    try:
        number = int(float(str(value).replace(",", "").strip()))
    except ValueError as exc:
        raise ValueError(f"{field} 숫자 형식 오류") from exc
    if number < 0:
        raise ValueError(f"{field}은 0 이상이어야 합니다")
    return number


def validate_measurement(total: int, ng: int, etc: int) -> None:
    if ng > total:
        raise ValueError("NG는 총체결보다 클 수 없습니다")
    if etc > total:
        raise ValueError("분류실패는 총체결보다 클 수 없습니다")


def rate(part: int, total: int) -> float:
    return 0 if total == 0 else round(part / total * 100, 2)


def previous_period(start: date, end: date) -> tuple[date, date]:
    days = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    return prev_end - timedelta(days=days - 1), prev_end


def measurement_dict(row: DailyMeasurement) -> dict[str, Any]:
    return {
        "id": row.id,
        "date": row.measurementDate.isoformat(),
        "line": row.process.line,
        "type": row.process.type,
        "processId": row.processId,
        "processName": row.process.processName,
        "status": row.process.status,
        "totalCount": row.totalCount,
        "ngCount": row.ngCount,
        "ngRate": rate(row.ngCount, row.totalCount),
        "etcCount": row.etcCount,
        "etcRate": rate(row.etcCount, row.totalCount),
        "clusterCount": row.clusterCount,
        "note": row.note or "",
        "updatedAt": row.updatedAt.isoformat(),
    }


def filtered_query(args: dict[str, Any]):
    query = DailyMeasurement.query.join(ProcessMaster)
    start = args.get("start")
    end = args.get("end")
    if start:
        query = query.filter(DailyMeasurement.measurementDate >= parse_date(start))
    if end:
        query = query.filter(DailyMeasurement.measurementDate <= parse_date(end))
    for key, col in {
        "line": ProcessMaster.line,
        "type": ProcessMaster.type,
        "process": ProcessMaster.processName,
        "status": ProcessMaster.status,
    }.items():
        values = args.getlist(key) if hasattr(args, "getlist") else args.get(key, [])
        if isinstance(values, str):
            values = [v for v in values.split(",") if v]
        if values:
            query = query.filter(col.in_(values))
    q = args.get("q")
    if q:
        like = f"%{q}%"
        query = query.filter(
            ProcessMaster.line.ilike(like)
            | ProcessMaster.type.ilike(like)
            | ProcessMaster.processName.ilike(like)
            | ProcessMaster.status.ilike(like)
            | DailyMeasurement.note.ilike(like)
        )
    return query


def summarize(query) -> dict[str, Any]:
    rows = query.all()
    total = sum(r.totalCount for r in rows)
    ng = sum(r.ngCount for r in rows)
    etc = sum(r.etcCount for r in rows)
    clusters = [r.clusterCount for r in rows]
    return {
        "totalCount": total,
        "ngCount": ng,
        "ngRate": rate(ng, total),
        "etcCount": etc,
        "etcRate": rate(etc, total),
        "avgCluster": round(sum(clusters) / len(clusters), 1) if clusters else 0,
        "processCount": len({r.processId for r in rows}),
        "dateCount": len({r.measurementDate for r in rows}),
    }


def add_audit(user, action: str, target: str, target_id: Any, before: Any, after: Any) -> None:
    db.session.add(
        AuditLog(
            userId=getattr(user, "id", None),
            username=getattr(user, "username", None),
            actionType=action,
            targetType=target,
            targetId=str(target_id) if target_id is not None else None,
            beforeValue=before,
            afterValue=after,
        )
    )


def dataframe_response(rows: list[dict[str, Any]], filename: str, fmt: str) -> Response:
    if fmt == "xlsx":
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        headers = list(rows[0].keys()) if rows else ["데이터"]
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
        wb.save(output)
        output.seek(0)
        return Response(
            output.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )
    output = io.StringIO()
    headers = list(rows[0].keys()) if rows else ["데이터"]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)
    return Response(
        "\ufeff" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
    )


def import_rows(raw_rows: list[dict[str, Any]], duplicate_mode: str, user) -> dict[str, Any]:
    rows = normalize_import(raw_rows)
    result = {"total": len(rows), "success": 0, "failed": 0, "duplicate": 0, "updated": 0, "created": 0, "errors": []}
    for idx, row in enumerate(rows, start=1):
        try:
            if row.get("__error"):
                raise ValueError(row["__error"])
            proc = ProcessMaster.query.filter_by(
                line=row["line"], type=row["type"], processName=row["processName"]
            ).first()
            if not proc:
                raise ValueError("존재하지 않는 공정")
            total, ng, etc = row["totalCount"], row["ngCount"], row["etcCount"]
            validate_measurement(total, ng, etc)
            existing = DailyMeasurement.query.filter_by(processId=proc.id, measurementDate=row["measurementDate"]).first()
            if existing:
                result["duplicate"] += 1
                if duplicate_mode == "keep" or duplicate_mode == "skip":
                    continue
                before = measurement_dict(existing)
                existing.totalCount = total
                existing.ngCount = ng
                existing.etcCount = etc
                existing.clusterCount = row["clusterCount"]
                existing.note = row["note"]
                existing.updatedBy = user.id
                result["updated"] += 1
                add_audit(user, "가져오기", "DailyMeasurement", existing.id, before, measurement_dict(existing))
            else:
                item = DailyMeasurement(processId=proc.id, createdBy=user.id, updatedBy=user.id, **row_measurement_only(row))
                db.session.add(item)
                db.session.flush()
                result["created"] += 1
                add_audit(user, "가져오기", "DailyMeasurement", item.id, None, measurement_dict(item))
            result["success"] += 1
        except Exception as exc:
            result["failed"] += 1
            result["errors"].append({**row, "row": idx, "reason": str(exc), "measurementDate": str(row.get("measurementDate", ""))})
    db.session.commit()
    return result


def row_measurement_only(row: dict[str, Any]) -> dict[str, Any]:
    return {k: row[k] for k in ["measurementDate", "totalCount", "ngCount", "etcCount", "clusterCount", "note"]}


def normalize_import(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cleaned = [{str(k).strip(): v for k, v in row.items()} for row in raw_rows]
    cols = set(cleaned[0].keys()) if cleaned else set()
    long_keys = {"날짜", "Line", "Type", "Process"}
    rows: list[dict[str, Any]] = []
    if long_keys.issubset(cols):
        for r in cleaned:
            base_row = {
                "line": str(r.get("Line", "")).strip(),
                "type": str(r.get("Type", "")).strip(),
                "processName": str(r.get("Process", "")).strip(),
                "measurementDate": r.get("날짜", ""),
                "totalCount": r.get("총체결", 0),
                "ngCount": r.get("NG", 0),
                "etcCount": r.get("분류실패", r.get("Etc", 0)),
                "clusterCount": r.get("Cluster", 0),
                "note": "" if r.get("비고", "") is None else str(r.get("비고", "")),
            }
            try:
                rows.append(
                    {
                        **base_row,
                        "measurementDate": parse_date(base_row["measurementDate"]),
                        "totalCount": to_int(base_row["totalCount"], "총체결"),
                        "ngCount": to_int(base_row["ngCount"], "NG"),
                        "etcCount": to_int(base_row["etcCount"], "분류실패"),
                        "clusterCount": to_int(base_row["clusterCount"], "Cluster"),
                    }
                )
            except Exception as exc:
                rows.append({**base_row, "__error": str(exc)})
        return rows

    base = ["Line", "Type", "Process", "현황", "구분"]
    if not set(base).issubset(cols):
        raise ValueError("Long Format 또는 Wide Format 필수 열이 없습니다")
    date_cols = [c for c in cleaned[0].keys() if c not in base]
    grouped: dict[tuple[str, str, str, date], dict[str, Any]] = {}
    for r in cleaned:
        for c in date_cols:
            try:
                measurement_date = parse_date(c)
            except Exception:
                continue
            key = (str(r["Line"]).strip(), str(r["Type"]).strip(), str(r["Process"]).strip(), measurement_date)
            item = grouped.setdefault(
                key,
                {
                    "line": key[0],
                    "type": key[1],
                    "processName": key[2],
                    "measurementDate": key[3],
                    "totalCount": 0,
                    "ngCount": 0,
                    "etcCount": 0,
                    "clusterCount": 0,
                    "note": "",
                },
            )
            label = str(r["구분"]).strip()
            val = r[c]
            if label == "총체결":
                item["totalCount"] = to_int(val, "총체결")
            elif label == "NG":
                item["ngCount"] = to_int(val, "NG")
            elif label in ["Etc", "Etc%", "분류실패", "분류실패%"]:
                if label in ["Etc", "분류실패"]:
                    item["etcCount"] = to_int(val, "분류실패")
            elif label == "Cluster":
                item["clusterCount"] = to_int(val, "Cluster")
            elif label == "비고":
                item["note"] = "" if val is None else str(val)
    return list(grouped.values())


def read_upload_rows(file) -> list[dict[str, Any]]:
    filename = file.filename.lower()
    if filename.endswith(".csv"):
        text = file.read().decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(c).strip() if c is not None else "" for c in values[0]]
    return [dict(zip(headers, row)) for row in values[1:] if any(v is not None for v in row)]
